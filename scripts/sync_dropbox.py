import json
import io
import os
import sys
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

REQUIRED_HEADERS = [
    "Country", "Area of check", "Source type", "Register / Portal", "Authority",
    "What to verify", "Search by", "Link", "Notes",
]


def download_url(shared_url: str) -> str:
    parsed = urllib.parse.urlparse(shared_url)
    query = dict(urllib.parse.parse_qsl(parsed.query))
    query["dl"] = "1"
    return urllib.parse.urlunparse(parsed._replace(query=urllib.parse.urlencode(query)))


def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


shared_url = os.environ.get("DROPBOX_SHARED_URL", "").strip()
if not shared_url:
    sys.exit("DROPBOX_SHARED_URL is empty")

with urllib.request.urlopen(download_url(shared_url), timeout=60) as response:
    workbook_bytes = response.read()

workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
worksheet = None
headers = []
for candidate in workbook.worksheets:
    candidate_headers = [cell_value(cell.value) for cell in next(candidate.iter_rows(min_row=1, max_row=1))]
    if all(header in candidate_headers for header in REQUIRED_HEADERS):
        worksheet = candidate
        headers = candidate_headers
        break

if worksheet is None:
    sys.exit("No worksheet contains all required Excel columns: " + ", ".join(REQUIRED_HEADERS))

indices = {header: headers.index(header) for header in REQUIRED_HEADERS}
rows = []
for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
    item = {header: cell_value(excel_row[index]) for header, index in indices.items()}
    if any(item.values()):
        rows.append(item)

if not rows:
    sys.exit("The workbook has no data rows")

generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
content = "window.REGISTRY_DATA = " + json.dumps(rows, ensure_ascii=False, indent=2) + ";\n"
content += "window.REGISTRY_DATA_UPDATED_AT = " + json.dumps(generated_at) + ";\n"
Path("data.js").write_text(content, encoding="utf-8")
print(f"Updated {len(rows)} rows from worksheet '{worksheet.title}' at {generated_at}")
